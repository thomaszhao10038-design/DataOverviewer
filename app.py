import streamlit as st
import pandas as pd
from io import BytesIO
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side, numbers
from openpyxl.chart import LineChart, Reference, Series
from openpyxl.utils import get_column_letter

# ==========================================
# PART 1: CONFIGURATION & HELPER FUNCTIONS
# ==========================================

st.set_page_config(
    page_title="Data Analyser Pro",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Constants
PSUM_OUTPUT_NAME = 'PSum (W)' 
POWER_COL_OUT = 'PSumW'
DATE_FORMAT_MAP = {
    "DD/MM/YYYY": "%d/%m/%Y %H:%M:%S",
    "YYYY-MM-DD": "%Y-%m-%d %H:%M:%S"
}

def excel_col_to_index(col_str):
    """Converts Excel column letter to 0-based index."""
    col_str = col_str.upper().strip()
    index = 0
    for char in col_str:
        if 'A' <= char <= 'Z':
            index = index * 26 + (ord(char) - ord('A') + 1)
        else:
            raise ValueError(f"Invalid character in column string: {col_str}")
    return index - 1

# ==========================================
# PART 2: CORE LOGIC (From app.py)
# ==========================================

def process_uploaded_files(uploaded_files, file_configs):
    """
    Reads multiple CSV files, extracts configured columns, cleans PSum data.
    """
    processed_data = {}
    
    for i, uploaded_file in enumerate(uploaded_files):
        filename = uploaded_file.name
        config = file_configs[i] 
        
        try:
            date_col_index = excel_col_to_index(config['date_col_str'])
            time_col_index = excel_col_to_index(config['time_col_str'])
            ps_um_col_index = excel_col_to_index(config['psum_col_str'])
            
            columns_to_extract = {
                date_col_index: 'Date',
                time_col_index: 'Time',
                ps_um_col_index: PSUM_OUTPUT_NAME
            }
            col_indices = list(columns_to_extract.keys())
            
            if len(set(col_indices)) != 3:
                st.error(f"Error for file **{filename}**: Columns must be unique.")
                continue
            
            header_index = int(config['start_row_num']) - 1 
            date_format_string = DATE_FORMAT_MAP.get(config['selected_date_format'])
            separator = config['delimiter_input']
            
            uploaded_file.seek(0)
            df_full = pd.read_csv(
                uploaded_file, 
                header=header_index, 
                encoding='ISO-8859-1', 
                low_memory=False,
                sep=separator
            )
            
            max_index = max(col_indices)
            if df_full.shape[1] < max_index + 1:
                st.error(f"File **{filename}** has insufficient columns. Check delimiter '{separator}'.")
                continue

            df_extracted = df_full.iloc[:, col_indices].copy()
            temp_cols = {k: v for k, v in columns_to_extract.items()}
            df_extracted.columns = temp_cols.values()
            
            if PSUM_OUTPUT_NAME in df_extracted.columns:
                df_extracted[PSUM_OUTPUT_NAME] = pd.to_numeric(
                    df_extracted[PSUM_OUTPUT_NAME].astype(str).str.replace(',', '.', regex=False), 
                    errors='coerce' 
                )

            combined_dt_str = df_extracted['Date'].astype(str) + ' ' + df_extracted['Time'].astype(str)
            datetime_series = pd.to_datetime(combined_dt_str, errors='coerce', format=date_format_string)
            
            if datetime_series.count() == 0:
                st.error(f"File **{filename}**: No valid dates could be parsed. Check settings.")
                continue

            df_final = pd.DataFrame({
                'Date': datetime_series.dt.strftime('%d/%m/%Y'), 
                'Time': datetime_series.dt.strftime('%H:%M:%S'),
                PSUM_OUTPUT_NAME: df_extracted[PSUM_OUTPUT_NAME] 
            })

            sheet_name = filename.replace('.csv', '').replace('.', '_').strip()[:31]
            processed_data[sheet_name] = df_final
            
        except ValueError as e:
            st.error(f"Configuration Error for file **{filename}**: {e}")
            continue
        except Exception as e:
            st.error(f"Error processing file **{filename}**: {e}")
            continue
            
    return processed_data

def process_sheet(df, date_col, time_col, psum_col):
    """
    Processes a single sheet: cleans, rounds to 10-min, filters zeros.
    """
    combined_dt_series = df[date_col].astype(str) + ' ' + df[time_col].astype(str)
    df['Timestamp'] = pd.to_datetime(combined_dt_series, errors="coerce", format='%d/%m/%Y %H:%M:%S')
    timestamp_col = 'Timestamp'
    
    power_series = df[psum_col].astype(str).str.strip().str.replace(',', '.', regex=False) 
    df[psum_col] = pd.to_numeric(power_series, errors='coerce')
    
    df = df.dropna(subset=[timestamp_col, psum_col])
    if df.empty: return pd.DataFrame()
    
    non_zero_indices = df[~np.isclose(df[psum_col].abs(), 0.0)].index
    if non_zero_indices.empty: return pd.DataFrame() 
        
    first_valid_idx = non_zero_indices.min()
    last_valid_idx = non_zero_indices.max()
    df = df.loc[first_valid_idx:last_valid_idx].copy()

    df_indexed = df.set_index(timestamp_col)
    df_indexed.index = df_indexed.index.floor('10min')
    resampled_data = df_indexed[psum_col].groupby(level=0).sum()
    
    df_out = resampled_data.reset_index()
    df_out.columns = ['Rounded', POWER_COL_OUT] 
    
    if df_out.empty or df_out['Rounded'].isna().all(): return pd.DataFrame()
    
    original_dates = set(df_out['Rounded'].dt.date)
    min_dt = df_out['Rounded'].min().floor('D')
    max_dt_exclusive = df_out['Rounded'].max().ceil('D') 
    
    if min_dt >= max_dt_exclusive: return pd.DataFrame()
    
    full_time_index = pd.date_range(
        start=min_dt.to_pydatetime(),
        end=max_dt_exclusive.to_pydatetime(),
        freq='10min',
        inclusive='left'
    )
    
    df_indexed_for_reindex = df_out.set_index('Rounded')
    df_padded_series = df_indexed_for_reindex[POWER_COL_OUT].reindex(full_time_index) 
    
    grouped = df_padded_series.reset_index().rename(columns={'index': 'Rounded'})
    grouped.columns = ['Rounded', POWER_COL_OUT]
    grouped[POWER_COL_OUT] = grouped[POWER_COL_OUT].astype(float) 

    grouped["Date"] = grouped["Rounded"].dt.date
    grouped["Time"] = grouped["Rounded"].dt.strftime("%H:%M") 
    grouped = grouped[grouped["Date"].isin(original_dates)]
    grouped['kW'] = grouped[POWER_COL_OUT].abs() / 1000
    
    return grouped

def prepare_chart_data(analysis_results):
    """
    Aggregates daily max kW for 'Total' sheet logic.
    """
    total_sheet_data = {}
    
    for sheet_name, df in analysis_results.items():
        df['Date'] = pd.to_datetime(df['Date']).dt.date
        daily_maxes = df.groupby('Date')['kW'].max().reset_index()
        
        for _, row in daily_maxes.iterrows():
            date = row['Date']
            max_kw = row['kW']
            if date not in total_sheet_data: total_sheet_data[date] = {}
            total_sheet_data[date][sheet_name] = max_kw

    dates = sorted(total_sheet_data.keys())
    chart_df = pd.DataFrame(index=dates)
    sheet_names_list = sorted(list(set(sheet for date_data in total_sheet_data.values() for sheet in date_data.keys())))
    
    for date in dates:
        for sheet_name in sheet_names_list:
            chart_df.loc[date, sheet_name] = total_sheet_data[date].get(sheet_name, 0)

    chart_df['Total Load (kW)'] = chart_df[sheet_names_list].sum(axis=1)
    chart_df.index.name = 'Date'
    
    return chart_df, total_sheet_data

def calculate_axis_interval(max_value):
    """Calculates a 'nice' axis interval (1, 2, 5, 10, etc.)."""
    if max_value <= 0: return 10
    num_intervals = 8
    raw_interval = max_value / num_intervals
    magnitude = 10 ** (int(np.log10(raw_interval)) if raw_interval > 0 else 0)
    normalized = raw_interval / magnitude
    
    if normalized <= 1.5: interval = 1 * magnitude
    elif normalized <= 3: interval = 2 * magnitude
    elif normalized <= 7: interval = 5 * magnitude
    else: interval = 10 * magnitude
    
    if interval < 1: interval = round(interval, 1)
    else: interval = int(round(interval))
    return max(interval, 1)

def build_output_excel(sheets_dict, total_sheet_data):
    """
    Generates the Excel file. Uses strict logic to ensure chart grid labels are nice numbers.
    """
    wb = Workbook()
    if 'Sheet' in wb.sheetnames: wb.remove(wb['Sheet'])

    header_fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')
    title_font = Font(bold=True, size=12)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Process Individual Sheets
    for sheet_name, df in sheets_dict.items():
        ws = wb.create_sheet(sheet_name)
        df['Date'] = pd.to_datetime(df['Date']).dt.date
        dates = sorted(df["Date"].unique())
        col_start = 1
        max_row_used = 0
        daily_max_summary = []
        day_intervals = []
        
        for date in dates:
            day_data_full = df[df["Date"] == date].sort_values("Time")
            day_data_active = day_data_full.dropna(subset=[POWER_COL_OUT])
            n_rows = len(day_data_full) 
            day_intervals.append(n_rows)
            
            data_start_row = 3
            merge_start = data_start_row
            merge_end = merge_start + n_rows - 1
            date_str_full = date.strftime('%Y-%m-%d')
            date_str_short = date.strftime('%d-%b')

            ws.merge_cells(start_row=1, start_column=col_start, end_row=1, end_column=col_start+3)
            ws.cell(row=1, column=col_start, value=date_str_full).alignment = Alignment(horizontal="center")
            
            ws.cell(row=2, column=col_start, value="UTC Offset (minutes)")
            ws.cell(row=2, column=col_start+1, value="Local Time Stamp")
            ws.cell(row=2, column=col_start+2, value="Active Power (W)")
            ws.cell(row=2, column=col_start+3, value="kW")

            ws.merge_cells(start_row=merge_start, start_column=col_start, end_row=merge_end, end_column=col_start)
            ws.cell(row=merge_start, column=col_start, value=date_str_full).alignment = Alignment(horizontal="center", vertical="center")

            for idx, r in enumerate(day_data_full.itertuples(), start=merge_start):
                ws.cell(row=idx, column=col_start+1, value=r.Time)
                ws.cell(row=idx, column=col_start+2, value=getattr(r, POWER_COL_OUT) if not pd.isna(getattr(r, POWER_COL_OUT)) else None) 
                ws.cell(row=idx, column=col_start+3, value=r.kW if not pd.isna(r.kW) else None)

            stats_row_start = merge_end + 1
            sum_w = day_data_active[POWER_COL_OUT].sum(); mean_w = day_data_active[POWER_COL_OUT].mean(); max_w = day_data_active[POWER_COL_OUT].max()
            sum_kw = day_data_active['kW'].sum(); mean_kw = day_data_active['kW'].mean(); max_kw = day_data_active['kW'].max()

            ws.cell(row=stats_row_start, column=col_start+1, value="Total")
            ws.cell(row=stats_row_start, column=col_start+2, value=sum_w)
            ws.cell(row=stats_row_start, column=col_start+3, value=sum_kw)
            ws.cell(row=stats_row_start+1, column=col_start+1, value="Average")
            ws.cell(row=stats_row_start+1, column=col_start+2, value=mean_w)
            ws.cell(row=stats_row_start+1, column=col_start+3, value=mean_kw)
            ws.cell(row=stats_row_start+2, column=col_start+1, value="Max")
            ws.cell(row=stats_row_start+2, column=col_start+2, value=max_w)
            ws.cell(row=stats_row_start+2, column=col_start+3, value=max_kw)

            max_row_used = max(max_row_used, stats_row_start+2)
            daily_max_summary.append((date_str_short, max_kw)) 
            col_start += 4

        if dates:
            chart = LineChart()
            chart.title = f"Daily 10-Minute Absolute Power Profile - {sheet_name}"
            chart.y_axis.title = "kW"; chart.x_axis.title = "Time"
            chart.height = 12.5; chart.width = 23    
            max_rows = max(day_intervals)
            first_time_col = 2
            categories_ref = Reference(ws, min_col=first_time_col, min_row=3, max_row=2 + max_rows)
            col_start = 1
            for i, n_rows in enumerate(day_intervals):
                data_ref = Reference(ws, min_col=col_start+3, min_row=3, max_col=col_start+3, max_row=2+n_rows)
                date_title_str = dates[i].strftime('%d-%b')
                s = Series(values=data_ref, title=date_title_str)
                chart.series.append(s)
                col_start += 4
            chart.set_categories(categories_ref)
            chart.x_axis.delete = False; chart.y_axis.delete = False
            
            # --- NICE NUMBER & FORMAT LOGIC START ---
            max_kw_value = df['kW'].max() if not df.empty and 'kW' in df.columns else 100
            min_kw_value = df['kW'].min() if not df.empty and 'kW' in df.columns else 0
            
            # Add padding to data max to find the next "Nice" Interval
            padded_max = max_kw_value * 1.05
            y_interval = calculate_axis_interval(padded_max)
            
            # FORCE INTEGER LABELS ON AXIS
            chart.y_axis.numFmt = '0'
            
            if y_interval > 0:
                # Snap the max scale to the next interval multiple
                nice_max = np.ceil(padded_max / y_interval) * y_interval
                chart.y_axis.scaling.max = nice_max
                chart.y_axis.scaling.min = max(0, min_kw_value * 0.9) if min_kw_value >= 0 else min_kw_value * 1.1
                chart.y_axis.majorUnit = y_interval
                chart.y_axis.majorTickMark = 'cross'
                chart.y_axis.scaling.auto = False
            # --- NICE NUMBER & FORMAT LOGIC END ---
            
            ws.add_chart(chart, f'G{max_row_used+2}')

        if daily_max_summary:
            start_row = max_row_used + 5 
            ws.cell(row=start_row, column=1, value="Daily Max Power (kW) Summary").font = title_font
            ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=2)
            start_row += 1
            ws.cell(row=start_row, column=1, value="Day").fill = header_fill
            ws.cell(row=start_row, column=2, value="Max (kW)").fill = header_fill
            for d, (date_str, max_kw) in enumerate(daily_max_summary):
                row = start_row+1+d
                ws.cell(row=row, column=1, value=date_str)
                ws.cell(row=row, column=2, value=max_kw).number_format = numbers.FORMAT_NUMBER_00
            ws.column_dimensions['A'].width = 15; ws.column_dimensions['B'].width = 15

    # TOTAL SHEET
    if total_sheet_data:
        ws_total = wb.create_sheet("Total")
        sorted_dates = sorted(total_sheet_data.keys())
        all_sheet_names = sorted(list(set(sheet for date_data in total_sheet_data.values() for sheet in date_data.keys())))
        headers = ["Date"] + all_sheet_names + ["Total Load"]
        for col_idx, header_text in enumerate(headers, 1):
            cell = ws_total.cell(row=1, column=col_idx, value=header_text)
            cell.font = title_font; cell.fill = header_fill; cell.alignment = Alignment(horizontal="center"); cell.border = thin_border
            ws_total.column_dimensions[get_column_letter(col_idx)].width = 20

        for row_idx, date_obj in enumerate(sorted_dates, 2):
            date_cell = ws_total.cell(row=row_idx, column=1, value=date_obj.strftime('%Y-%m-%d'))
            date_cell.border = thin_border; date_cell.alignment = Alignment(horizontal="center")
            row_total_load = 0
            for col_idx, sheet_name in enumerate(all_sheet_names, 2):
                val = total_sheet_data[date_obj].get(sheet_name, 0)
                if pd.isna(val): val = 0
                cell = ws_total.cell(row=row_idx, column=col_idx, value=val)
                cell.number_format = numbers.FORMAT_NUMBER_00; cell.border = thin_border
                row_total_load += val
            total_cell = ws_total.cell(row=row_idx, column=len(all_sheet_names) + 2, value=row_total_load)
            total_cell.number_format = numbers.FORMAT_NUMBER_00; total_cell.border = thin_border; total_cell.font = Font(bold=True)

        if sorted_dates:
            chart_total = LineChart()
            chart_total.title = "Daily Max Load Overview" 
            chart_total.y_axis.title = "Max Power (kW)"; chart_total.x_axis.title = "Date"
            chart_total.height = 15; chart_total.width = 30
            data_max_row = len(sorted_dates) + 1
            total_cols = len(all_sheet_names) + 2
            
            data_ref = Reference(ws_total, min_col=2, min_row=1, max_col=total_cols, max_row=data_max_row)
            chart_total.add_data(data_ref, titles_from_data=True)

            # Ensure Straight Lines
            for s in chart_total.series:
                s.smooth = False
            
            cats_ref = Reference(ws_total, min_col=1, min_row=2, max_row=data_max_row)
            chart_total.set_categories(cats_ref)
            chart_total.x_axis.delete = False; chart_total.y_axis.delete = False
            
            # --- NICE NUMBER & FORMAT LOGIC START (TOTAL SHEET) ---
            max_total_value = 0; min_total_value = float('inf')
            for date_data in total_sheet_data.values():
                for sheet_name in all_sheet_names:
                    val = date_data.get(sheet_name, 0)
                    if pd.notna(val): max_total_value = max(max_total_value, val); min_total_value = min(min_total_value, val)
            for date_obj in sorted_dates:
                row_total = sum(total_sheet_data[date_obj].get(sheet_name, 0) for sheet_name in all_sheet_names)
                max_total_value = max(max_total_value, row_total); min_total_value = min(min_total_value, row_total)
            if min_total_value == float('inf'): min_total_value = 0
            
            padded_max = max_total_value * 1.05
            y_interval = calculate_axis_interval(padded_max)
            
            # FORCE INTEGER LABELS ON AXIS
            chart_total.y_axis.numFmt = '0'

            if y_interval > 0:
                nice_max = np.ceil(padded_max / y_interval) * y_interval
                chart_total.y_axis.scaling.max = nice_max
                chart_total.y_axis.scaling.min = max(0, min_total_value * 0.9) if min_total_value >= 0 else min_total_value * 1.1
                chart_total.y_axis.majorUnit = y_interval
                chart_total.y_axis.majorTickMark = 'cross'
                chart_total.y_axis.scaling.auto = False
            # --- NICE NUMBER & FORMAT LOGIC END ---
            
            ws_total.add_chart(chart_total, "B" + str(data_max_row + 3))

    stream = BytesIO()
    if 'Sheet' in wb.sheetnames and len(wb.sheetnames) > len(sheets_dict) + (1 if total_sheet_data else 0): wb.remove(wb['Sheet'])
    wb.save(stream)
    stream.seek(0)
    return stream

# ==========================================
# PART 3: MAIN APP INTERFACE (UX OVERHAUL)
# ==========================================

st.title("⚡Data Analyser Pro for MSB")
st.markdown("""
1. Upload multiple raw CSV files (up to 10).
2. Click **Generate Final Report** to get the analyzed Excel file.
""")

uploaded_files = st.file_uploader("Choose CSV files (Max 10)", type=["csv"], accept_multiple_files=True)
file_configs = []

if uploaded_files:
    if len(uploaded_files) > 10:
        st.warning(f"Only first 10 files will be processed."); uploaded_files = uploaded_files[:10]
    
    # --- Sidebar Configuration (Matches Good UI:UX style) ---
    with st.sidebar:
        st.header("Configuration")
        st.write("Configure settings for each file:")
        for i, uploaded_file in enumerate(uploaded_files):
            with st.expander(f"⚙️ Settings: {uploaded_file.name}", expanded=(i==0)):
                col1, col2 = st.columns(2)
                with col1:
                    date_col = st.text_input("Date Column Letter", value='A', key=f'd_{i}')
                    time_col = st.text_input("Time Column Letter", value='B', key=f't_{i}')
                    psum_col = st.text_input("PSum Column Letter", value='BI', key=f'p_{i}', help="Active Power column")
                with col2:
                    sep = st.text_input("Delimiter", value=',', key=f'sep_{i}')
                    row = st.number_input("Header Row", min_value=1, value=3, key=f'row_{i}')
                    fmt = st.selectbox("Date Format", ["DD/MM/YYYY", "YYYY-MM-DD"], key=f'fmt_{i}')
                
                # Map back to original keys expected by process_uploaded_files
                file_configs.append({
                    'date_col_str': date_col, 'time_col_str': time_col, 'psum_col_str': psum_col,
                    'delimiter_input': sep, 'start_row_num': row, 'selected_date_format': fmt
                })

    if st.button("🚀 Generate Final Report"):
        st.info("Reading and extracting CSV data...")
        
        # 1. Extraction (Old Stage 1)
        raw_dfs = process_uploaded_files(uploaded_files, file_configs)
        
        if len(raw_dfs) != len(uploaded_files):
            st.error(f"Failed: {len(uploaded_files) - len(raw_dfs)} file(s) could not be parsed. Check sidebar settings.")
            st.stop()
        
        st.success(f"Extracted data from {len(raw_dfs)} files.")
        st.info("Performing analysis and generating charts...")
        
        # 2. Analysis (Old Stage 2)
        analyzed_sheets = {}
        for sheet_name, df in raw_dfs.items():
            analyzed_df = process_sheet(df, 'Date', 'Time', PSUM_OUTPUT_NAME)
            if not analyzed_df.empty:
                analyzed_sheets[sheet_name] = analyzed_df
        
        if analyzed_sheets:
            # 3. Chart Prep & Excel Build (Preserved Logic)
            chart_df, total_data = prepare_chart_data(analyzed_sheets)
            
            # Note: Screen chart removed as per request.
            # Directly generate Excel.
            excel_file = build_output_excel(analyzed_sheets, total_data)
            
            st.write("---")
            st.subheader("✅ Result Ready")
            st.download_button(
                label="📥 Download Final Energy Report (.xlsx)",
                data=excel_file,
                file_name="Final_Energy_Report.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary"
            )
        else:
            st.error("Analysis failed. No valid power readings found in any file.")
else:
    st.info("Awaiting file upload.")
